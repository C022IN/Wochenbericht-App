"""
Import corrected KW 18–22 Excel data into Supabase wochenbericht_entries.

Files read:
  wb-kw18/AXIANS OFM Wochenbericht April 2026 KW 18.xlsx
  wb-kw18/AXIANS OFM Wochenbericht Mai 2026 KW 18.xlsx
  wb-kw19-22/AXIANS OFM Wochenbericht Mai 2026 KW 19.xlsx
  wb-kw19-22/AXIANS OFM Wochenbericht Mai 2026 KW 20.xlsx
  wb-kw19-22/AXIANS OFM Wochenbericht Mai 2026 KW 21.xlsx
  wb-kw19-22/AXIANS OFM Wochenbericht Mai 2026 KW 22.xlsx

Run:  python scripts/import_corrected_kw18_22.py
"""

import json
import os
import re
import uuid
from datetime import datetime, time, date
from pathlib import Path

import openpyxl
import requests

# ---------------------------------------------------------------------------
# Load env vars from .env.local
# ---------------------------------------------------------------------------
ENV_FILE = Path(__file__).parent.parent / ".env.local"
env: dict[str, str] = {}
with open(ENV_FILE, encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")

SUPABASE_URL = env["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
ENTRIES_TABLE = env.get("SUPABASE_ENTRIES_TABLE", "wochenbericht_entries")
PROFILES_TABLE = env.get("SUPABASE_PROFILES_TABLE", "wochenbericht_profiles")

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

# ---------------------------------------------------------------------------
# Resolve user ID from profiles table
# ---------------------------------------------------------------------------
TARGET_EMAIL = "collin-ambani.anjeo@axians.de"


def get_user_id() -> str:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/{PROFILES_TABLE}?select=user_id,email&limit=10",
        headers=HEADERS,
        timeout=10,
    )
    resp.raise_for_status()
    rows = resp.json()
    if not rows:
        raise RuntimeError("No profiles found in Supabase — is the DB seeded?")
    print(f"  Found {len(rows)} profile(s):")
    for r in rows:
        print(f"    {r['user_id']}  email={r.get('email', '?')}")
    # Prefer the row matching the target email
    for r in rows:
        if r.get("email", "").strip().lower() == TARGET_EMAIL.lower():
            print(f"  Selected by email match: {r['user_id']}")
            return r["user_id"]
    # Fallback: first row
    print(f"  No email match for {TARGET_EMAIL!r}; using first profile")
    return rows[0]["user_id"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
WEEKDAY_COLS = ["H", "I", "J", "K", "L", "M", "N"]  # Mon..Sun


def fmt_time(v) -> str:
    """Convert openpyxl time cell to HH:MM string."""
    if v is None:
        return ""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return ""


def fmt_num(v) -> str:
    """Convert numeric/string cell to a clean decimal string (German comma → dot)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, (int, float)):
        # Remove unnecessary trailing zeros: 2.5 → "2.5", 8.0 → "8"
        s = f"{v:.4f}".rstrip("0").rstrip(".")
        return s if s else ""
    s = str(v).strip().replace(",", ".")
    return s


def parse_von_date(v) -> tuple[int, int]:
    """Return (year, month) from the 'von' cell (L1)."""
    if v is None:
        raise ValueError("L1 (von date) is empty")
    if isinstance(v, (datetime, date)):
        return v.year, v.month
    s = str(v).strip()
    # German format: "27.04.2026" or "27.04.26" (2-digit year)
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return y, int(m.group(2))
    # ISO: "2026-04-27"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    raise ValueError(f"Cannot parse von date: {s!r}")


def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# Read one xlsx file → dict[iso_date, DailyEntry-dict]
# ---------------------------------------------------------------------------
def read_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Wochenbericht"]

    # Header info
    year, month = parse_von_date(ws["L1"].value)
    arbeitsstaette = clean_str(ws["D5"].value)
    art_der_arbeit = clean_str(ws["D6"].value)

    # Day-of-month headers from row 9 (H9..N9)
    day_of_month: dict[str, int] = {}
    for col in WEEKDAY_COLS:
        raw = clean_str(ws[f"{col}9"].value)
        stripped = raw.rstrip("*").strip()
        if stripped.isdigit():
            day_of_month[col] = int(stripped)

    # Collect lines grouped by ISO date
    by_date: dict[str, list[dict]] = {}

    for row_no in range(10, 50):
        site = clean_str(ws[f"A{row_no}"].value)

        # Find which weekday column has a value
        target_col: str | None = None
        hours_raw = None
        for col in WEEKDAY_COLS:
            cell_val = ws[f"{col}{row_no}"].value
            if cell_val is not None:
                target_col = col
                hours_raw = cell_val
                break

        # Skip rows with no day marker and no meaningful data
        if target_col is None:
            continue
        if target_col not in day_of_month:
            continue  # column exists in week but day header was empty (split-month edge)

        dom = day_of_month[target_col]
        iso_date = f"{year}-{month:02d}-{dom:02d}"

        # Parse fields
        beginn = fmt_time(ws[f"E{row_no}"].value)
        ende = fmt_time(ws[f"F{row_no}"].value)
        pause = fmt_num(ws[f"G{row_no}"].value)
        lohn_type = clean_str(ws[f"Q{row_no}"].value)
        ausloese = clean_str(ws[f"R{row_no}"].value)
        zulage = fmt_num(ws[f"S{row_no}"].value)
        projektnummer = clean_str(ws[f"T{row_no}"].value)
        kabelschacht = clean_str(ws[f"U{row_no}"].value)
        sm_nr = clean_str(ws[f"V{row_no}"].value)
        bauleiter = clean_str(ws[f"W{row_no}"].value)
        kollege = clean_str(ws[f"X{row_no}"].value)

        # Determine hours / lineType
        if isinstance(hours_raw, bool):
            hours_str = ""
            line_type = "arbeitszeit"
        elif isinstance(hours_raw, str):
            s = hours_raw.strip().replace(",", ".")
            hours_str = s
            line_type = "baustelle" if s.lower() == "x" else "arbeitszeit"
        elif isinstance(hours_raw, (int, float)):
            hours_str = fmt_num(hours_raw)
            line_type = "arbeitszeit"
        else:
            hours_str = ""
            line_type = "arbeitszeit"

        # Skip completely empty rows
        has_data = any([site, projektnummer, bauleiter, kollege, hours_str, beginn, ende])
        if not has_data:
            continue

        line = {
            "id": str(uuid.uuid4()),
            "lineType": line_type,
            "siteNameOrt": site,
            "beginn": beginn,
            "ende": ende,
            "pauseOverride": pause,
            "dayHoursOverride": hours_str,
            "lohnType": lohn_type if lohn_type else "S",
            "ausloese": ausloese,
            "zulage": zulage,
            "projektnummer": projektnummer,
            "kabelschachtInfo": kabelschacht,
            "smNr": sm_nr,
            "bauleiter": bauleiter,
            "arbeitskollege": kollege,
        }

        by_date.setdefault(iso_date, []).append(line)

    # Build DailyEntry objects
    entries: dict[str, dict] = {}
    for iso_date, lines in by_date.items():
        entries[iso_date] = {
            "date": iso_date,
            "arbeitsstaetteProjekte": arbeitsstaette,
            "artDerArbeit": art_der_arbeit,
            "lines": lines,
            "updatedAt": datetime.utcnow().isoformat() + "Z",
        }

    return entries


# ---------------------------------------------------------------------------
# Upsert one DailyEntry into Supabase
# ---------------------------------------------------------------------------
def upsert_entry(user_id: str, entry: dict) -> None:
    iso_date = entry["date"]
    payload = {
        "user_id": user_id,
        "date": iso_date,
        "payload": entry,
        "updated_at": entry["updatedAt"],
    }
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/{ENTRIES_TABLE}?on_conflict=user_id,date",
        headers={**HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"},
        json=payload,
        timeout=10,
    )
    if resp.status_code not in (200, 201, 204):
        raise RuntimeError(f"Upsert failed for {iso_date}: {resp.status_code} {resp.text}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
FILES = [
    Path(r"C:\Users\Window\Downloads\wb-kw18\AXIANS OFM Wochenbericht April 2026 KW 18.xlsx"),
    Path(r"C:\Users\Window\Downloads\wb-kw18\AXIANS OFM Wochenbericht Mai 2026 KW 18.xlsx"),
    Path(r"C:\Users\Window\Downloads\wb-kw19-22\AXIANS OFM Wochenbericht Mai 2026 KW 19.xlsx"),
    Path(r"C:\Users\Window\Downloads\wb-kw19-22\AXIANS OFM Wochenbericht Mai 2026 KW 20.xlsx"),
    Path(r"C:\Users\Window\Downloads\wb-kw19-22\AXIANS OFM Wochenbericht Mai 2026 KW 21.xlsx"),
    Path(r"C:\Users\Window\Downloads\wb-kw19-22\AXIANS OFM Wochenbericht Mai 2026 KW 22.xlsx"),
]


def main() -> None:
    print("Resolving user ID from Supabase profiles...")
    user_id = get_user_id()
    print(f"  user_id = {user_id}")
    print()

    all_entries: dict[str, dict] = {}

    for xlsx_path in FILES:
        if not xlsx_path.exists():
            print(f"  SKIP (not found): {xlsx_path.name}")
            continue
        print(f"Reading {xlsx_path.name} ...")
        entries = read_xlsx(xlsx_path)
        print(f"  {len(entries)} day(s): {sorted(entries.keys())}")
        # Merge (later files win for same date — shouldn't happen across our set)
        all_entries.update(entries)

    print()
    print(f"Total days to upsert: {len(all_entries)}")
    print("Upserting...")

    for iso_date in sorted(all_entries.keys()):
        entry = all_entries[iso_date]
        upsert_entry(user_id, entry)
        n_lines = len(entry["lines"])
        print(f"  OK {iso_date}  ({n_lines} line{'s' if n_lines != 1 else ''})")

    print()
    print("Done. All entries imported successfully.")


if __name__ == "__main__":
    main()
