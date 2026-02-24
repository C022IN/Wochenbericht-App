import argparse
import json
import math
import os
import warnings
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported*")
warnings.filterwarnings("ignore", message="Data Validation extension is not supported*")
warnings.filterwarnings("ignore", message="wmf image format is not supported*")


WEEKDAY_COLUMNS = ["H", "I", "J", "K", "L", "M", "N"]
DATA_ROW_START = 10
DATA_ROW_END = 49


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload-file", required=True, help="JSON payload file")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    return parser.parse_args()


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_time_value(value: str):
    if not value or not isinstance(value, str):
        return None
    try:
        hh, mm = value.split(":")
        return time(hour=int(hh), minute=int(mm))
    except Exception:
        return None


def parse_decimal(value: str):
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    txt = value.strip().replace(",", ".")
    if not txt:
        return None
    try:
        num = float(txt)
        if math.isfinite(num):
            return num
    except ValueError:
        return value.strip()
    return value.strip()


def gross_hours(start: time | None, end: time | None):
    if start is None or end is None:
        return None
    start_minutes = start.hour * 60 + start.minute
    end_minutes = end.hour * 60 + end.minute
    diff = end_minutes - start_minutes
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


def auto_pause_hours(hours: float | None):
    if hours is None:
        return None
    if hours > 9.5:
        return 0.75
    if hours > 6:
        return 0.5
    return 0.0


def infer_pause_from_net_hours(net_hours: float | None):
    if net_hours is None:
        return None
    # Ambiguous edge cases exist when only net hours are known. Prefer the smallest
    # pause that is consistent with the template's hardcoded auto-pause thresholds.
    for pause in (0.0, 0.5, 0.75):
        gross = net_hours + pause
        if auto_pause_hours(gross) == pause:
            return pause
    return None


def compute_day_cell_value(row: dict):
    override = row.get("dayHoursOverride")
    if isinstance(override, str):
        override_str = override.strip()
        if override_str and override_str != "__AUTO_FROM_TIME__":
            return parse_decimal(override_str)
    elif override is not None:
        return override

    start = parse_time_value(row.get("beginn", ""))
    end = parse_time_value(row.get("ende", ""))
    gross = gross_hours(start, end)
    if gross is None:
        return None

    pause_override = parse_decimal(row.get("pauseOverride", ""))
    if isinstance(pause_override, (int, float)):
        return round(gross - float(pause_override), 2)

    pause_auto = auto_pause_hours(gross) or 0.0
    return round(gross - pause_auto, 2)


def write_header(ws, payload):
    ws["H1"] = int(payload["kw"])
    ws["L1"] = payload["reportStartDe"]  # Template uses text formatting in L1.
    ws["R1"] = parse_iso_date(payload["reportEnd"])

    profile = payload["profile"]
    ws["D3"] = profile.get("name", "")
    ws["P3"] = profile.get("vorname", "")
    ws["D5"] = profile.get("arbeitsstaetteProjekte", "")
    ws["D6"] = profile.get("artDerArbeit", "")


def clear_and_write_date_row(ws, payload):
    segment_dates = set(payload["segmentDates"])
    all_week_dates = payload["allWeekDates"]

    for col in WEEKDAY_COLUMNS:
        ws[f"{col}9"] = None

    for iso in all_week_dates:
        if iso not in segment_dates:
            continue
        d = parse_iso_date(iso)
        iso_weekday = d.isoweekday()  # 1..7
        col = WEEKDAY_COLUMNS[iso_weekday - 1]
        ws[f"{col}9"] = d.day


def clear_data_rows(ws):
    for row in range(DATA_ROW_START, DATA_ROW_END + 1):
        ws[f"A{row}"] = None
        ws[f"E{row}"] = None
        ws[f"F{row}"] = None
        for col in WEEKDAY_COLUMNS:
            ws[f"{col}{row}"] = None
        for col in ["Q", "R", "S", "T", "U", "V", "W", "X"]:
            ws[f"{col}{row}"] = None
        # Keep template formulas in G/O/P intact unless pause override is explicitly written.


def write_rows(ws, payload):
    rows = payload.get("rows", [])
    max_rows = DATA_ROW_END - DATA_ROW_START + 1
    truncated = max(0, len(rows) - max_rows)
    rows_to_write = rows[:max_rows]

    for idx, row_data in enumerate(rows_to_write):
        row_no = DATA_ROW_START + idx
        iso = row_data.get("date")
        day_cell_value = compute_day_cell_value(row_data)
        weekday_col = None
        if isinstance(iso, str):
            try:
                weekday_col = WEEKDAY_COLUMNS[parse_iso_date(iso).isoweekday() - 1]
            except Exception:
                weekday_col = None

        ws[f"A{row_no}"] = row_data.get("siteNameOrt", "")
        start_t = parse_time_value(row_data.get("beginn", ""))
        end_t = parse_time_value(row_data.get("ende", ""))
        if start_t:
            ws[f"E{row_no}"] = start_t
        if end_t:
            ws[f"F{row_no}"] = end_t

        pause_override = parse_decimal(row_data.get("pauseOverride", ""))
        if isinstance(pause_override, (int, float)):
            ws[f"G{row_no}"] = float(pause_override)
        elif not start_t and not end_t and isinstance(day_cell_value, (int, float)):
            pause_auto_from_hours = infer_pause_from_net_hours(float(day_cell_value))
            if isinstance(pause_auto_from_hours, (int, float)) and pause_auto_from_hours > 0:
                ws[f"G{row_no}"] = float(pause_auto_from_hours)

        if weekday_col and isinstance(day_cell_value, (int, float)) and day_cell_value >= 0:
            ws[f"{weekday_col}{row_no}"] = float(day_cell_value)
        elif weekday_col and isinstance(day_cell_value, str) and day_cell_value.strip():
            marker = day_cell_value.strip()
            ws[f"{weekday_col}{row_no}"] = "x" if marker.lower() == "x" else marker

        ws[f"Q{row_no}"] = row_data.get("lohnType", "")
        ws[f"R{row_no}"] = row_data.get("ausloese", "")
        zulage = parse_decimal(row_data.get("zulage", ""))
        ws[f"S{row_no}"] = zulage if zulage is not None else ""
        ws[f"T{row_no}"] = row_data.get("projektnummer", "")
        ws[f"U{row_no}"] = row_data.get("kabelschachtInfo", "")

        sm_nr = row_data.get("smNr", "")
        sm_num = parse_decimal(sm_nr) if isinstance(sm_nr, str) else sm_nr
        ws[f"V{row_no}"] = sm_num if sm_num is not None else ""

        ws[f"W{row_no}"] = row_data.get("bauleiter", "")
        ws[f"X{row_no}"] = row_data.get("arbeitskollege", "")

    return len(rows_to_write), truncated


def main():
    args = parse_args()
    payload_path = Path(args.payload_file)
    output_path = Path(args.output)

    data = json.loads(payload_path.read_text(encoding="utf-8"))
    template_path = data["templatePath"]
    payload = data["payload"]

    wb = load_workbook(template_path)
    if "Wochenbericht" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Wochenbericht' not found in template")
    ws = wb["Wochenbericht"]

    write_header(ws, payload)
    clear_and_write_date_row(ws, payload)
    clear_data_rows(ws)
    rows_written, rows_truncated = write_rows(ws, payload)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    result = {
      "output_path": str(output_path),
      "rows_written": rows_written,
      "rows_truncated": rows_truncated,
      "warnings": []
    }
    if rows_truncated:
        result["warnings"].append(
            f"More than 40 lines for this report. Export truncated by {rows_truncated} line(s) to fit Excel rows 10-49."
        )

    print(json.dumps(result))


if __name__ == "__main__":
    main()
